from openpyxl import Workbook, load_workbook
import os


def open_and_write_excel(filepath: str, row_data: list, sheet_name="Gefahrstoffkataster", insert_row: int | None = None):
    """
    Append or insert a row into an Excel file. Creates file if it does not exist.

    :param filepath: path to the .xlsx file
    :param row_data: list of values to write
    :param sheet_name: name of the sheet
    :param insert_row: if given, insert at this row (1-based index).
                       If None, append at the end.
    """

    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.create_sheet(sheet_name)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
        # Write header row on first creation
        ws.append([
            "Produktname / Handelsname",
            "Hersteller",
            "UN-Nr.",
            "Gefahren (H-Sätze)",
            "Piktogramme",
            "Lagerort",
            "Menge im Lager",
            "Besonderheiten",
            "SDS"
            "Stand"
        ])

    if insert_row is not None:
        # Ensure at least 1 (headers are row 1)
        insert_row = max(2, insert_row)
        ws.insert_rows(insert_row)
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=insert_row, column=col_idx, value=value)
    else:
        ws.append(row_data)

    wb.save(filepath)


def convert_data_to_list(data: dict) -> list:
    """
    Converts the SDS data to a list for writing to Excel.
    :param data: the dict of SDS data
    :return: list of values
    """
    return [
        data["handelsname"],
        data["manufacturer"],
        data["un_number"],
        ", ".join(data["h_statements"]),
        ", ".join(data["pictograms"]),
        "",  # Lagerort
        "",  # Menge im Lager
        "",  # Besonderheiten
        "",  # SDS Path
        data["sds_date"],
    ]
